"""Tests for renaming names in place: the plan, and the three rewriters."""

from __future__ import annotations

import json

import pytest

from pytoolbox.dataset import edit, writers
from pytoolbox.dataset.errors import DataError

NAMES = ("id", "First Name", "age")


def test_matches_a_name_exactly():
    plan = edit.plan(NAMES, ["id=user_id"])
    assert [(item.index, item.old, item.new) for item in plan.renames] == [(0, "id", "user_id")]


def test_matches_a_name_whatever_its_case():
    plan = edit.plan(NAMES, ["first name=full_name"])
    assert [item.old for item in plan.renames] == ["First Name"]


def test_matches_the_sql_spelling_of_a_name():
    plan = edit.plan(NAMES, ["first_name=full_name"])
    assert [item.old for item in plan.renames] == ["First Name"]


def test_keeps_the_new_name_verbatim():
    plan = edit.plan(NAMES, ["id=User ID"])
    assert plan.renames[0].new == "User ID"


def test_unknown_name_names_a_real_one():
    with pytest.raises(DataError) as excinfo:
        edit.plan(NAMES, ["ige=years"])
    assert "Did you mean: age?" in str(excinfo.value)


def test_unknown_name_with_nothing_like_it_lists_the_columns():
    with pytest.raises(DataError) as excinfo:
        edit.plan(NAMES, ["zzzz=years"])
    assert "Columns: id, First Name, age." in str(excinfo.value)


@pytest.mark.parametrize("pair", ["id", "id=", "=user_id", "  =  "])
def test_a_malformed_pair_is_refused(pair):
    with pytest.raises(DataError, match="wants OLD=NEW"):
        edit.plan(NAMES, [pair])


def test_two_columns_renamed_to_the_same_name_are_refused():
    with pytest.raises(DataError, match="collide"):
        edit.plan(NAMES, ["id=who", "age=who"])


def test_renaming_onto_an_untouched_column_is_refused():
    with pytest.raises(DataError, match="already called 'age'"):
        edit.plan(NAMES, ["id=age"])


def test_swapping_two_names_is_allowed():
    plan = edit.plan(NAMES, ["id=age", "age=id"])
    assert plan.applied() == ("age", "First Name", "id")


def test_renaming_a_name_to_itself_is_nothing_to_do():
    assert not edit.plan(NAMES, ["id=id"])


def test_the_last_word_wins_when_a_column_is_named_twice():
    plan = edit.plan(NAMES, ["id=first", "id=second"])
    assert [item.new for item in plan.renames] == ["second"]


# --- rewriting CSV ---------------------------------------------------------

def rows(text):
    return text.splitlines(keepends=True)[1:]


def test_csv_header_is_renamed():
    text = "id,name\n1,ann\n"
    out = writers.rewrite_csv(text, edit.plan(("id", "name"), ["name=full_name"]), ",")
    assert out.splitlines()[0] == "id,full_name"


def test_csv_data_rows_are_untouched():
    text = 'id,name\n1,"ann, jr"\n2,  bob  \n'
    out = writers.rewrite_csv(text, edit.plan(("id", "name"), ["id=user_id"]), ",")
    assert rows(out) == rows(text)


def test_csv_keeps_its_delimiter():
    text = "id;name\n1;ann\n"
    out = writers.rewrite_csv(text, edit.plan(("id", "name"), ["id=user_id"]), ";")
    assert out.splitlines()[0] == "user_id;name"


def test_csv_keeps_crlf_line_endings():
    text = "id,name\r\n1,ann\r\n"
    out = writers.rewrite_csv(text, edit.plan(("id", "name"), ["id=user_id"]), ",")
    assert out == "user_id,name\r\n1,ann\r\n"


def test_csv_header_spanning_two_lines_is_spliced_back_correctly():
    text = 'id,"two\nline"\n1,ann\n'
    out = writers.rewrite_csv(text, edit.plan(("id", "two\nline"), ["id=user_id"]), ",")
    assert out == 'user_id,"two\nline"\n1,ann\n'


def test_a_new_name_containing_the_delimiter_is_quoted():
    text = "id,name\n1,ann\n"
    out = writers.rewrite_csv(text, edit.plan(("id", "name"), ["name=last, first"]), ",")
    assert out.splitlines()[0] == 'id,"last, first"'


# --- rewriting JSON --------------------------------------------------------

def test_json_key_is_renamed_in_every_record():
    text = json.dumps([{"id": 1}, {"id": 2}])
    out = writers.rewrite_json(text, edit.plan(("id",), ["id=user_id"]))
    assert json.loads(out) == [{"user_id": 1}, {"user_id": 2}]


def test_json_keeps_the_order_of_the_keys():
    text = json.dumps([{"a": 1, "id": 2, "z": 3}])
    out = writers.rewrite_json(text, edit.plan(("a", "id", "z"), ["id=user_id"]))
    assert list(json.loads(out)[0]) == ["a", "user_id", "z"]


def test_json_leaves_the_envelope_around_root_alone():
    text = json.dumps({"meta": {"page": 1}, "data": {"users": [{"id": 1}]}})
    out = writers.rewrite_json(text, edit.plan(("id",), ["id=user_id"]), root="data.users")
    assert json.loads(out) == {"meta": {"page": 1}, "data": {"users": [{"user_id": 1}]}}


def test_json_keeps_the_indentation_it_was_written_with():
    text = json.dumps([{"id": 1}], indent=4)
    out = writers.rewrite_json(text, edit.plan(("id",), ["id=user_id"]))
    assert '\n    {\n        "user_id": 1' in out


def test_json_written_on_one_line_stays_on_one_line():
    text = '[{"id": 1}]'
    out = writers.rewrite_json(text, edit.plan(("id",), ["id=user_id"]))
    assert out.strip() == '[{"user_id": 1}]'


def test_json_keeps_its_trailing_newline():
    out = writers.rewrite_json('[{"id": 1}]\n', edit.plan(("id",), ["id=user_id"]))
    assert out.endswith("}]\n")


def test_json_without_a_trailing_newline_gains_none():
    out = writers.rewrite_json('[{"id": 1}]', edit.plan(("id",), ["id=user_id"]))
    assert not out.endswith("\n")


def test_json_keeps_non_ascii_keys_readable():
    text = json.dumps([{"نام": "ann"}], ensure_ascii=False)
    out = writers.rewrite_json(text, edit.plan(("نام",), ["نام=نام واحد"]))
    assert '"نام واحد"' in out


def test_ndjson_stays_one_object_per_line():
    text = '{"id": 1}\n{"id": 2}\n'
    out = writers.rewrite_json(text, edit.plan(("id",), ["id=user_id"]))
    assert out == '{"user_id": 1}\n{"user_id": 2}\n'


def test_json_single_object_is_renamed():
    out = writers.rewrite_json('{"id": 1}', edit.plan(("id",), ["id=user_id"]))
    assert json.loads(out) == {"user_id": 1}


# --- writing to disk -------------------------------------------------------

def test_apply_writes_the_file_and_a_backup(tmp_path):
    path = tmp_path / "sales.csv"
    path.write_text("id,name\n1,ann\n", encoding="utf-8")
    backup = writers.apply(edit.plan(("id", "name"), ["id=user_id"]), path, "csv", delimiter=",")
    assert path.read_text(encoding="utf-8") == "user_id,name\n1,ann\n"
    assert backup.read_text(encoding="utf-8") == "id,name\n1,ann\n"


def test_apply_can_skip_the_backup(tmp_path):
    path = tmp_path / "sales.csv"
    path.write_text("id,name\n1,ann\n", encoding="utf-8")
    assert writers.apply(edit.plan(("id",), ["id=user_id"]), path, "csv", delimiter=",", backup=False) is None
    assert not (tmp_path / "sales.csv.bak").exists()


def test_apply_to_a_target_leaves_the_original_alone(tmp_path):
    path = tmp_path / "sales.csv"
    path.write_text("id,name\n1,ann\n", encoding="utf-8")
    target = tmp_path / "renamed.csv"
    writers.apply(edit.plan(("id",), ["id=user_id"]), path, "csv", target=target, delimiter=",")
    assert path.read_text(encoding="utf-8") == "id,name\n1,ann\n"
    assert target.read_text(encoding="utf-8") == "user_id,name\n1,ann\n"


def test_a_failed_write_leaves_no_scratch_file_behind(tmp_path, monkeypatch):
    path = tmp_path / "sales.csv"
    path.write_text("id,name\n1,ann\n", encoding="utf-8")

    def explode(*args, **kwargs):
        raise OSError("disk full")

    monkeypatch.setattr(writers.os, "replace", explode)
    with pytest.raises(OSError):
        writers.apply(edit.plan(("id",), ["id=user_id"]), path, "csv", delimiter=",", backup=False)
    assert [item.name for item in tmp_path.iterdir()] == ["sales.csv"]
    assert path.read_text(encoding="utf-8") == "id,name\n1,ann\n"


# --- Excel -----------------------------------------------------------------

@pytest.fixture
def workbook(tmp_path):
    pytest.importorskip("openpyxl", reason="Excel support is an optional extra")
    from openpyxl import Workbook
    from openpyxl.styles import Font

    book = Workbook()
    sheet = book.active
    sheet.title = "Staff"
    sheet.append(["ID", "Full Name", "Salary"])
    sheet.append([1, "ann", 100])
    sheet.append([2, "bob", 200])
    sheet["D2"] = "=C2*2"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.column_dimensions["B"].width = 42
    book.create_sheet("Notes").append(["kept"])
    path = tmp_path / "staff.xlsx"
    book.save(path)
    return path


def test_excel_header_cell_is_renamed(workbook, tmp_path):
    from openpyxl import load_workbook

    writers.apply(edit.plan(("ID", "Full Name", "Salary"), ["Full Name=name"]), workbook, "excel")
    sheet = load_workbook(workbook).active
    assert [cell.value for cell in sheet[1]][:3] == ["ID", "name", "Salary"]


def test_excel_keeps_formulas_styles_widths_and_other_sheets(workbook):
    from openpyxl import load_workbook

    writers.apply(edit.plan(("ID", "Full Name", "Salary"), ["ID=id"]), workbook, "excel")
    book = load_workbook(workbook)
    sheet = book["Staff"]
    assert sheet["D2"].value == "=C2*2"
    assert sheet["A1"].font.bold and sheet["A1"].font.size == 14
    assert sheet.column_dimensions["B"].width == 42
    assert book["Notes"]["A1"].value == "kept"
    assert [cell.value for cell in sheet[2]] == [1, "ann", 100, "=C2*2"]


def test_excel_renames_on_the_named_sheet(workbook):
    from openpyxl import load_workbook

    writers.apply(edit.plan(("kept",), ["kept=notes"]), workbook, "excel", sheet="Notes")
    book = load_workbook(workbook)
    assert book["Notes"]["A1"].value == "notes"
    assert book["Staff"]["A1"].value == "ID"


def test_an_unknown_sheet_is_refused(workbook):
    with pytest.raises(DataError, match="No sheet named 'Q9'"):
        writers.apply(edit.plan(("ID",), ["ID=id"]), workbook, "excel", sheet="Q9")
