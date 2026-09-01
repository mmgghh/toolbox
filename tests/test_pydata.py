"""Tests for the pydata CLI: the four subcommands, end to end."""

from __future__ import annotations

import datetime as dt
import json
import sqlite3

import pytest

from pytoolbox.pydata import data_cli

openpyxl = pytest.importorskip("openpyxl", reason="Excel support is an optional extra")

USERS = [
    {"id": 1, "First Name": "ann", "age": 34, "tags": ["a", "b"], "active": True},
    {"id": 2, "First Name": "bob", "age": 28.5, "tags": [], "active": False, "note": "hi"},
    {"id": 3, "First Name": "cy", "age": None, "tags": ["c"], "active": True},
]


@pytest.fixture
def api(tmp_path):
    path = tmp_path / "api.json"
    path.write_text(json.dumps({"meta": {"page": 1}, "data": {"users": USERS}}), encoding="utf-8")
    return path


@pytest.fixture
def flat(tmp_path):
    path = tmp_path / "users.json"
    path.write_text(json.dumps(USERS), encoding="utf-8")
    return path


@pytest.fixture
def sales(tmp_path):
    path = tmp_path / "sales.csv"
    path.write_text(
        "id,name,joined,zip,amount\n1,ann,2024-01-05,01730,120.50\n2,bob,2024-03-11,10115,7\n",
        encoding="utf-8",
    )
    return path


@pytest.fixture
def orders(tmp_path):
    path = tmp_path / "orders.json"
    path.write_text(
        json.dumps(
            [
                {
                    "id": 1,
                    "address": {"city": "Berlin", "zip": "10115"},
                    "items": [{"sku": "a", "city": "X"}, {"sku": "b"}],
                },
                {"id": 2, "address": {"city": "Rome"}, "items": []},
            ]
        ),
        encoding="utf-8",
    )
    return path


@pytest.fixture
def multi_sheet(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    q1 = wb.active
    q1.title = "Q1"
    q1.append(["ID", "Amount"])
    q1.append([1, 10])
    q1.append([2, 20])
    q2 = wb.create_sheet("Q2")
    q2.append(["ID", "Amount"])
    q2.append([3, 30])
    notes = wb.create_sheet("Notes")
    notes.append(["ID", "Text"])
    notes.append([1, "hi"])
    path = tmp_path / "multi.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def book(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Staff"
    sheet.append(["ID", "Full Name", "Hired"])
    sheet.append([1, "ann", dt.datetime(2024, 1, 5)])
    sheet.append([2, "bob", None])
    path = tmp_path / "staff.xlsx"
    wb.save(path)
    return path


def run(runner, *args, **kwargs):
    return runner.invoke(data_cli, [str(arg) for arg in args], **kwargs)


# ── tree ────────────────────────────────────────────────────────────


def test_tree_finds_the_records_inside_an_envelope(runner, api):
    result = run(runner, "tree", api)
    assert result.exit_code == 0
    assert "--root data.users" in result.stdout
    assert "list of 3 objects" in result.stdout
    assert "list[str]" in result.stdout


def test_tree_reports_the_root_it_chose_on_stderr(runner, api):
    result = run(runner, "tree", api)
    assert "Using --root data.users" in result.stderr


def test_tree_reads_a_csv(runner, sales):
    result = run(runner, "tree", sales)
    assert result.exit_code == 0
    assert "joined" in result.stdout
    assert "date" in result.stdout


def test_tree_reads_an_xlsx(runner, book):
    result = run(runner, "tree", book)
    assert result.exit_code == 0
    assert "Full Name" in result.stdout


def test_tree_reads_stdin_with_from(runner):
    result = run(runner, "tree", "-", "--from", "json", input='[{"a": 1}]')
    assert result.exit_code == 0
    assert "stdin" in result.stdout


def test_tree_depth_limits_the_descent(runner, tmp_path):
    path = tmp_path / "a.json"
    path.write_text(json.dumps([{"a": {"b": {"c": 1}}}]), encoding="utf-8")
    output = run(runner, "tree", path, "--depth", "2").stdout
    assert "-- b" in output
    assert "-- c" not in output


# ── summary ─────────────────────────────────────────────────────────


def test_summary_prints_a_row_per_field(runner, api):
    result = run(runner, "summary", api)
    assert result.exit_code == 0
    assert "First Name" in result.stdout
    assert "first_name" in result.stdout
    assert "non_null" in result.stdout


def test_summary_as_json_stays_parseable(runner, api):
    result = run(runner, "summary", api, "--format", "json")
    assert result.exit_code == 0
    rows = json.loads(result.stdout)
    assert [row["field"] for row in rows][:2] == ["id", "First Name"]


def test_summary_writes_a_file(runner, api, tmp_path):
    out = tmp_path / "s.csv"
    result = run(runner, "summary", api, "--format", "csv", "-o", out)
    assert result.exit_code == 0
    assert "field,column,type" in out.read_text()


# ── filter ──────────────────────────────────────────────────────────


def test_filter_by_type(runner, api):
    result = run(runner, "filter", api, "--type", "int", "--type", "float")
    assert result.exit_code == 0
    assert "id" in result.stdout
    assert "age" in result.stdout
    assert "first_name" not in result.stdout


def test_filter_by_key_glob(runner, api):
    result = run(runner, "filter", api, "-k", "first*")
    assert "first_name" in result.stdout
    assert "tags" not in result.stdout


def test_filter_renders_containers_as_json(runner, api):
    assert '["a","b"]' in run(runner, "filter", api, "-k", "tags").stdout


def test_filter_rows_caps_the_output(runner, api):
    result = run(runner, "filter", api, "-k", "id", "--rows", "2")
    assert "3" not in result.stdout


def test_filter_with_no_match_fails_clearly(runner, api):
    result = run(runner, "filter", api, "-k", "zzz")
    assert result.exit_code == 1
    assert "No field matched" in result.stderr


def test_filter_deep_finds_a_key_at_any_depth(runner, orders):
    result = run(runner, "filter", orders, "-k", "city", "--deep")
    assert result.exit_code == 0
    assert "address.city" in result.stdout
    assert "items.[].city" in result.stdout
    assert "Berlin" in result.stdout
    assert "Rome" in result.stdout
    assert "X" in result.stdout


def test_filter_deep_reports_the_record_number(runner, orders):
    result = run(runner, "filter", orders, "-k", "city", "--deep", "--format", "json")
    rows = json.loads(result.stdout)
    assert [row["record"] for row in rows] == [1, 1, 2]


def test_filter_deep_drop_empty_skips_null_values(runner, tmp_path):
    path = tmp_path / "a.json"
    path.write_text(json.dumps([{"a": {"b": None}}, {"a": {"b": 1}}]), encoding="utf-8")
    result = run(runner, "filter", path, "-k", "b", "--deep", "--drop-empty", "--format", "json")
    rows = json.loads(result.stdout)
    assert [row["value"] for row in rows] == [1]


def test_filter_deep_rejects_type_mixed(runner, orders):
    result = run(runner, "filter", orders, "--deep", "--type", "mixed")
    assert result.exit_code == 1
    assert "--deep" in result.stderr


def test_filter_deep_with_no_match_fails_clearly(runner, orders):
    result = run(runner, "filter", orders, "-k", "zzz", "--deep")
    assert result.exit_code == 1
    assert "No field matched" in result.stderr


def test_filter_sheet_wildcard_merges_matching_sheets(runner, multi_sheet):
    result = run(runner, "filter", multi_sheet, "-k", "amount", "--sheet", "*")
    assert result.exit_code == 0
    assert "Q1" in result.stdout and "Q2" in result.stdout
    assert "Notes" not in result.stdout
    assert result.stdout.count("10") == 1 and "30" in result.stdout


def test_filter_sheet_wildcard_skips_a_sheet_without_the_field(runner, multi_sheet):
    result = run(runner, "filter", multi_sheet, "-k", "text", "--sheet", "*", "--format", "json")
    rows = json.loads(result.stdout)
    assert [row["sheet"] for row in rows] == ["Notes"]


def test_filter_sheet_wildcard_rejects_non_excel(runner, sales):
    result = run(runner, "filter", sales, "-k", "id", "--sheet", "*")
    assert result.exit_code == 1
    assert "only applies to Excel" in result.stderr


def test_filter_sheet_wildcard_no_match_anywhere(runner, multi_sheet):
    result = run(runner, "filter", multi_sheet, "-k", "zzz", "--sheet", "*")
    assert result.exit_code == 1
    assert "No field matched the filter in any sheet" in result.stderr


def test_filter_sheet_wildcard_with_deep(runner, multi_sheet):
    result = run(runner, "filter", multi_sheet, "-k", "text", "--deep", "--sheet", "*", "--format", "json")
    rows = json.loads(result.stdout)
    assert rows == [{"sheet": "Notes", "record": 1, "path": "Text", "value": "hi"}]


# ── count ───────────────────────────────────────────────────────────


def test_count_a_json_envelope(runner, api):
    assert run(runner, "count", api).stdout.strip() == "3 records"


def test_count_respects_root(runner, api):
    assert run(runner, "count", api, "--root", "data.users").stdout.strip() == "3 records"


def test_count_a_csv(runner, sales):
    assert run(runner, "count", sales).stdout.strip() == "2 records"


def test_count_stdin_with_from(runner):
    result = run(runner, "count", "-", "--from", "json", input='[{"a": 1}]')
    assert result.stdout.strip() == "1 record"


def test_count_an_xlsx_sheet_explicitly(runner, book):
    assert run(runner, "count", book, "--sheet", "Staff").stdout.strip() == "2 records"


def test_count_an_xlsx_without_sheet_lists_every_sheet(runner, tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    staff = wb.active
    staff.title = "Staff"
    staff.append(["ID", "Name"])
    staff.append([1, "ann"])
    staff.append([2, "bob"])
    empty = wb.create_sheet("Empty")
    empty.append(["ID"])
    path = tmp_path / "book.xlsx"
    wb.save(path)

    result = run(runner, "count", path)
    assert result.exit_code == 0
    assert "Staff" in result.stdout and "2" in result.stdout
    assert "Empty" in result.stdout and "0" in result.stdout


# ── keys ────────────────────────────────────────────────────────────


def test_keys_a_json_envelope(runner, api):
    result = run(runner, "keys", api)
    assert result.exit_code == 0
    assert "first_name" in result.stdout
    assert "First Name" not in result.stdout


def test_keys_raw_names_keeps_the_original_spelling(runner, api):
    result = run(runner, "keys", api, "--raw-names")
    assert "First Name" in result.stdout


def test_keys_a_csv(runner, sales):
    result = run(runner, "keys", sales)
    assert result.exit_code == 0
    assert "joined" in result.stdout
    assert "amount" in result.stdout


def test_keys_an_xlsx_sheet_explicitly(runner, book):
    result = run(runner, "keys", book, "--sheet", "Staff")
    assert result.exit_code == 0
    assert "full_name" in result.stdout


def test_keys_an_xlsx_without_sheet_lists_every_sheet(runner, tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    staff = wb.active
    staff.title = "Staff"
    staff.append(["ID", "Full Name"])
    staff.append([1, "ann"])
    empty = wb.create_sheet("Empty")
    empty.append(["Only Header"])
    path = tmp_path / "book.xlsx"
    wb.save(path)

    result = run(runner, "keys", path)
    assert result.exit_code == 0
    assert "Staff" in result.stdout and "full_name" in result.stdout
    assert "Empty" in result.stdout and "only_header" in result.stdout


def test_keys_groups_sheets_that_share_the_same_columns(runner, tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    first = wb.active
    first.title = "Q1"
    first.append(["ID", "Amount"])
    first.append([1, 10])
    second = wb.create_sheet("Q2")
    second.append(["ID", "Amount"])
    second.append([1, 20])
    other = wb.create_sheet("Other")
    other.append(["Name"])
    other.append(["x"])
    path = tmp_path / "book.xlsx"
    wb.save(path)

    result = run(runner, "keys", path)
    assert result.exit_code == 0
    assert result.stdout.count("id") == 1
    assert result.stdout.count("amount") == 1
    assert "Q1, Q2:" in result.stdout
    assert "Other:" in result.stdout


def test_keys_as_json_stays_parseable(runner, sales):
    result = run(runner, "keys", sales, "--format", "json")
    assert result.exit_code == 0
    rows = json.loads(result.stdout)
    assert [row["key"] for row in rows] == ["id", "name", "joined", "zip", "amount"]


# ── sql: scripts ────────────────────────────────────────────────────


def test_sql_writes_a_postgres_script_to_stdout(runner, api):
    result = run(runner, "sql", api, "-t", "users", "--dialect", "postgres", "--sql", "-")
    assert result.exit_code == 0
    assert 'CREATE TABLE "users"' in result.stdout
    assert "jsonb" in result.stdout
    assert "TRUE" in result.stdout


def test_sql_defaults_to_sqlite(runner, api):
    result = run(runner, "sql", api, "-t", "users", "--sql", "-")
    assert "INTEGER" in result.stdout
    assert "jsonb" not in result.stdout


def test_sql_writes_a_script_file(runner, api, tmp_path):
    out = tmp_path / "users.sql"
    result = run(runner, "sql", api, "-t", "users", "--sql", out)
    assert result.exit_code == 0
    assert "CREATE TABLE" in out.read_text()
    assert "SQL written to" in result.stderr


def test_sql_renames_a_column(runner, api):
    result = run(runner, "sql", api, "-t", "u", "-c", "first_name=full_name", "--sql", "-")
    assert '"full_name"' in result.stdout
    assert '"first_name"' not in result.stdout


def test_sql_adds_a_primary_key_and_indexes(runner, api):
    result = run(
        runner, "sql", api, "-t", "u", "--pk", "id", "--index", "active", "--unique-index", "first_name",
        "--sql", "-",
    )
    assert 'PRIMARY KEY ("id")' in result.stdout
    assert 'CREATE INDEX "idx_u_active"' in result.stdout
    assert 'CREATE UNIQUE INDEX "idx_u_first_name"' in result.stdout


def test_sql_supports_a_compound_primary_key(runner, api):
    result = run(runner, "sql", api, "-t", "u", "--pk", "id", "--pk", "first_name", "--sql", "-")
    assert 'PRIMARY KEY ("id", "first_name")' in result.stdout


def test_sql_supports_a_composite_index(runner, api):
    result = run(runner, "sql", api, "-t", "u", "--index", "id,active", "--sql", "-")
    assert '("id", "active")' in result.stdout


def test_sql_batches_inserts(runner, api):
    result = run(runner, "sql", api, "-t", "u", "--batch", "1", "--sql", "-")
    assert result.stdout.count("INSERT INTO") == 3


def test_dry_run_prints_without_writing(runner, api, tmp_path):
    db = tmp_path / "app.db"
    result = run(runner, "sql", api, "-t", "u", "--db", db, "--dry-run")
    assert result.exit_code == 0
    assert "CREATE TABLE" in result.stdout
    assert not db.exists()


# ── non-Latin headers ───────────────────────────────────────────────


PERSIAN_HEADERS = [
    "نام واحد",
    "ای دی واحد",
    "نوع ساختمان",
    "انبار/سردخانه",
    "ظرفیت",
    "کد پستی",
    "تاریخ ایجاد",
    "استان",
]


@pytest.fixture
def persian(tmp_path):
    """A spreadsheet whose header row is entirely Persian."""
    from openpyxl import Workbook

    wb = Workbook()
    sheet = wb.active
    sheet.title = "واحدها"
    sheet.append(PERSIAN_HEADERS)
    sheet.append(["انبار مرکزی", 1042, "فلزی", "انبار", 1200, "1968815453", dt.datetime(2024, 1, 5), "تهران"])
    sheet.append(["سردخانه شمال", 1043, "بتنی", "سردخانه", 800, "4671234567", dt.datetime(2023, 11, 2), "گیلان"])
    path = tmp_path / "vahed.xlsx"
    wb.save(path)
    return path


def test_persian_headers_are_not_flattened_to_column_1(runner, persian):
    """Every non-Latin header used to sanitize away to the same fallback name."""
    result = run(runner, "summary", persian, "--format", "json")
    assert result.exit_code == 0
    columns = [row["column"] for row in json.loads(result.stdout)]
    assert "column" not in columns
    assert not any(name.startswith("column_") for name in columns)
    assert len(set(columns)) == len(PERSIAN_HEADERS)


def test_persian_headers_keep_their_letters(runner, persian):
    rows = {row["field"]: row["column"] for row in json.loads(
        run(runner, "summary", persian, "--format", "json").stdout
    )}
    assert rows["نام واحد"] == "نام_واحد"
    assert rows["انبار/سردخانه"] == "انبار_سردخانه"
    assert rows["ظرفیت"] == "ظرفیت"


def test_a_persian_table_reaches_sqlite(runner, persian, tmp_path):
    db = tmp_path / "v.db"
    result = run(runner, "sql", persian, "-t", "واحدها", "--pk", "ای_دی_واحد", "--db", db)
    assert result.exit_code == 0

    connection = sqlite3.connect(db)
    names = [row[1] for row in connection.execute('PRAGMA table_info("واحدها")')]
    rows = connection.execute(
        'SELECT "نام_واحد", "استان", "ظرفیت" FROM "واحدها" ORDER BY "ای_دی_واحد"'
    ).fetchall()
    connection.close()
    assert len(names) == len(PERSIAN_HEADERS)
    assert rows == [("انبار مرکزی", "تهران", 1200), ("سردخانه شمال", "گیلان", 800)]


def test_a_persian_postal_code_stays_text(runner, persian):
    """A leading-zero-free numeric string is still an identifier, not a number."""
    rows = {row["field"]: row["type"] for row in json.loads(
        run(runner, "summary", persian, "--format", "json").stdout
    )}
    assert rows["کد پستی"] == "str"


def test_the_tree_shows_persian_keys(runner, persian):
    assert "نام واحد" in run(runner, "tree", persian).stdout


# ── sql: --raw-names ────────────────────────────────────────────────


def test_raw_names_keeps_the_keys_as_columns(runner, api):
    result = run(runner, "sql", api, "-t", "users", "--dialect", "postgres", "--raw-names", "--sql", "-")
    assert '"First Name" text' in result.stdout
    assert "first_name" not in result.stdout


def test_folding_is_still_the_default(runner, api):
    result = run(runner, "sql", api, "-t", "users", "--dialect", "postgres", "--sql", "-")
    assert '"first_name" text' in result.stdout
    assert '"First Name"' not in result.stdout


def test_raw_names_reach_sqlite(runner, api, tmp_path):
    db = tmp_path / "app.db"
    run(runner, "sql", api, "-t", "users", "--raw-names", "--db", db)
    connection = sqlite3.connect(db)
    names = [row[1] for row in connection.execute("PRAGMA table_info(users)")]
    value = connection.execute('SELECT "First Name" FROM users LIMIT 1').fetchone()[0]
    connection.close()
    assert "First Name" in names
    assert value == "ann"


def test_raw_names_honours_an_explicit_rename_verbatim(runner, api):
    result = run(runner, "sql", api, "-t", "u", "--raw-names", "-c", "First Name=Full Name", "--sql", "-")
    assert '"Full Name"' in result.stdout


def test_raw_names_applies_to_summary(runner, api):
    result = run(runner, "summary", api, "--raw-names", "--format", "json")
    rows = {row["field"]: row["column"] for row in json.loads(result.stdout)}
    assert rows["First Name"] == "First Name"


def test_raw_names_applies_to_filter(runner, api):
    result = run(runner, "filter", api, "--raw-names", "-k", "First*")
    assert "First Name" in result.stdout


def test_a_pk_under_raw_names_uses_the_raw_column(runner, api):
    result = run(runner, "sql", api, "-t", "u", "--raw-names", "--pk", "First Name", "--sql", "-")
    assert 'PRIMARY KEY ("First Name")' in result.stdout


# ── sql: SQLite ─────────────────────────────────────────────────────


def test_sql_fills_a_sqlite_file(runner, api, tmp_path):
    db = tmp_path / "app.db"
    result = run(runner, "sql", api, "-t", "users", "--pk", "id", "--db", db)
    assert result.exit_code == 0
    assert "Wrote 3 rows" in result.stderr

    connection = sqlite3.connect(db)
    rows = connection.execute("SELECT id, first_name, tags, note FROM users ORDER BY id").fetchall()
    connection.close()
    assert rows[0] == (1, "ann", '["a","b"]', None)
    assert rows[1][3] == "hi"


def test_a_missing_key_becomes_a_nullable_column(runner, api, tmp_path):
    """'note' is present in one record out of three."""
    db = tmp_path / "app.db"
    run(runner, "sql", api, "-t", "users", "--db", db)
    connection = sqlite3.connect(db)
    notnull = {row[1]: row[3] for row in connection.execute("PRAGMA table_info(users)")}
    connection.close()
    assert notnull["note"] == 0
    assert notnull["id"] == 1


def test_csv_types_reach_sqlite(runner, sales, tmp_path):
    db = tmp_path / "app.db"
    run(runner, "sql", sales, "-t", "sales", "--db", db)
    connection = sqlite3.connect(db)
    types = {row[1]: row[2] for row in connection.execute("PRAGMA table_info(sales)")}
    zip_value = connection.execute("SELECT zip FROM sales LIMIT 1").fetchone()[0]
    connection.close()
    assert types == {"id": "INTEGER", "name": "TEXT", "joined": "TEXT", "zip": "TEXT", "amount": "REAL"}
    assert zip_value == "01730"


def test_excel_reaches_sqlite(runner, book, tmp_path):
    db = tmp_path / "app.db"
    run(runner, "sql", book, "-t", "staff", "--db", db)
    connection = sqlite3.connect(db)
    rows = connection.execute("SELECT id, full_name, hired FROM staff ORDER BY id").fetchall()
    connection.close()
    assert rows == [(1, "ann", "2024-01-05T00:00:00"), (2, "bob", None)]


def test_a_single_object_becomes_one_row(runner, tmp_path):
    path = tmp_path / "cfg.json"
    path.write_text(json.dumps({"name": "cfg", "opts": {"x": 1}}), encoding="utf-8")
    db = tmp_path / "app.db"
    run(runner, "sql", path, "-t", "cfg", "--db", db)
    connection = sqlite3.connect(db)
    rows = connection.execute("SELECT name, opts FROM cfg").fetchall()
    connection.close()
    assert rows == [("cfg", '{"x":1}')]


def test_if_exists_replace_rebuilds(runner, api, tmp_path):
    db = tmp_path / "app.db"
    run(runner, "sql", api, "-t", "u", "--db", db)
    result = run(runner, "sql", api, "-t", "u", "--db", db, "--if-exists", "replace")
    assert result.exit_code == 0


def test_if_exists_append_adds_rows(runner, api, tmp_path):
    db = tmp_path / "app.db"
    run(runner, "sql", api, "-t", "u", "--db", db)
    run(runner, "sql", api, "-t", "u", "--db", db, "--if-exists", "append")
    connection = sqlite3.connect(db)
    count = connection.execute("SELECT count(*) FROM u").fetchone()[0]
    connection.close()
    assert count == 6


# ── sql: refusals ───────────────────────────────────────────────────


def test_an_existing_table_is_refused_by_default(runner, api, tmp_path):
    db = tmp_path / "app.db"
    run(runner, "sql", api, "-t", "u", "--db", db)
    result = run(runner, "sql", api, "-t", "u", "--db", db)
    assert result.exit_code == 1
    assert "already exists" in result.stderr


def test_db_and_sql_together_are_refused(runner, api, tmp_path):
    result = run(runner, "sql", api, "-t", "u", "--db", tmp_path / "a.db", "--sql", tmp_path / "a.sql")
    assert result.exit_code == 1
    assert "not both" in result.stderr


def test_postgres_cannot_be_written_directly(runner, api, tmp_path):
    result = run(runner, "sql", api, "-t", "u", "--dialect", "postgres", "--db", tmp_path / "a.db")
    assert result.exit_code == 1
    assert "generate a script" in result.stderr


def test_a_table_name_is_required(runner, api):
    result = run(runner, "sql", api, "--sql", "-")
    assert result.exit_code == 1
    assert "table name is required" in result.stderr


def test_nested_on_sqlite_warns_that_it_does_nothing(runner, api):
    result = run(runner, "sql", api, "-t", "u", "--nested", "text", "--sql", "-")
    assert result.exit_code == 0
    assert "SQLite has no JSON column type" in result.stderr


def test_an_ambiguous_document_names_the_candidates(runner, tmp_path):
    path = tmp_path / "two.json"
    path.write_text(json.dumps({"users": [{"a": 1}], "orders": [{"b": 2}]}), encoding="utf-8")
    result = run(runner, "tree", path)
    assert result.exit_code == 1
    assert "--root users" in result.stderr
    assert "--root orders" in result.stderr


# ── sql: interactive ────────────────────────────────────────────────


def test_interactive_asks_for_the_table_key_and_indexes(runner, flat):
    answers = "people\n\nid\nactive\n\ny\n"
    result = run(runner, "sql", flat, "-i", "--dialect", "postgres", "--sql", "-", input=answers)
    assert result.exit_code == 0
    assert 'CREATE TABLE "people"' in result.stdout
    assert 'PRIMARY KEY ("id")' in result.stdout
    assert 'CREATE INDEX "idx_people_active"' in result.stdout


def test_interactive_shows_the_summary_first(runner, flat):
    result = run(runner, "sql", flat, "-i", "--sql", "-", input="t\n\n\n\n\ny\n")
    assert "non_null" in result.stderr
    assert "distinct" in result.stderr


def test_interactive_defaults_the_table_name_to_the_filename(runner, flat):
    result = run(runner, "sql", flat, "-i", "--sql", "-", input="\n\n\n\n\ny\n")
    assert 'CREATE TABLE "users"' in result.stdout


def test_interactive_suggests_only_usable_keys(runner, flat):
    result = run(runner, "sql", flat, "-i", "--sql", "-", input="t\n\n\n\n\ny\n")
    suggestions = [line for line in result.stderr.splitlines() if "unique and complete" in line][0]
    assert "id" in suggestions
    assert "tags" not in suggestions


def test_interactive_re_asks_after_an_unknown_column(runner, flat):
    result = run(runner, "sql", flat, "-i", "--sql", "-", input="t\n\nnope\nid\n\n\ny\n")
    assert "No column called 'nope'" in result.stderr
    assert 'PRIMARY KEY ("id")' in result.stdout


def test_declining_at_the_end_writes_nothing(runner, flat, tmp_path):
    db = tmp_path / "app.db"
    result = run(runner, "sql", flat, "-i", "--db", db, input="t\n\n\n\n\nn\n")
    assert result.exit_code == 1
    assert not db.exists()


def test_interactive_composite_index(runner, flat):
    result = run(runner, "sql", flat, "-i", "--sql", "-", input="t\n\n\nid+active\n\ny\n")
    assert '("id", "active")' in result.stdout


# ── edit ────────────────────────────────────────────────────────────


def test_edit_renames_a_csv_header(runner, sales):
    result = run(runner, "edit", sales, "--rename", "name=full_name", "-y")
    assert result.exit_code == 0
    assert sales.read_text(encoding="utf-8").splitlines()[0] == "id,full_name,joined,zip,amount"


def test_edit_leaves_the_data_rows_of_a_csv_alone(runner, sales):
    before = sales.read_text(encoding="utf-8").splitlines(keepends=True)[1:]
    run(runner, "edit", sales, "--rename", "name=full_name", "-y")
    assert sales.read_text(encoding="utf-8").splitlines(keepends=True)[1:] == before


def test_edit_renames_a_json_key_inside_an_envelope(runner, api):
    result = run(runner, "edit", api, "--root", "data.users", "--rename", "First Name=full_name", "-y")
    assert result.exit_code == 0
    document = json.loads(api.read_text(encoding="utf-8"))
    assert document["meta"] == {"page": 1}
    assert list(document["data"]["users"][0]) == ["id", "full_name", "age", "tags", "active"]


def test_edit_renames_an_xlsx_header(runner, book):
    result = run(runner, "edit", book, "--rename", "Full Name=name", "-y")
    assert result.exit_code == 0
    sheet = openpyxl.load_workbook(book).active
    assert [cell.value for cell in sheet[1]] == ["ID", "name", "Hired"]


def test_edit_accepts_the_sql_spelling_of_a_name(runner, book):
    result = run(runner, "edit", book, "--rename", "full_name=name", "-y")
    assert result.exit_code == 0
    assert openpyxl.load_workbook(book).active["B1"].value == "name"


def test_edit_takes_the_column_spelling_sql_uses(runner, sales):
    result = run(runner, "edit", sales, "-c", "name=full_name", "-y")
    assert result.exit_code == 0
    assert "full_name" in sales.read_text(encoding="utf-8")


def test_edit_prints_what_it_changed(runner, sales):
    result = run(runner, "edit", sales, "--rename", "name=full_name", "-y")
    assert "name" in result.stdout and "full_name" in result.stdout


def test_edit_writes_a_backup(runner, sales):
    original = sales.read_text(encoding="utf-8")
    run(runner, "edit", sales, "--rename", "name=full_name", "-y")
    assert (sales.parent / "sales.csv.bak").read_text(encoding="utf-8") == original


def test_edit_can_skip_the_backup(runner, sales):
    run(runner, "edit", sales, "--rename", "name=full_name", "--no-backup", "-y")
    assert not (sales.parent / "sales.csv.bak").exists()


def test_edit_to_an_output_file_leaves_the_source_alone(runner, sales, tmp_path):
    target = tmp_path / "clean.csv"
    run(runner, "edit", sales, "--rename", "name=full_name", "-o", target, "-y")
    assert "full_name" in target.read_text(encoding="utf-8")
    assert "full_name" not in sales.read_text(encoding="utf-8")
    assert not (sales.parent / "sales.csv.bak").exists()


def test_edit_dry_run_writes_nothing(runner, sales):
    original = sales.read_text(encoding="utf-8")
    result = run(runner, "edit", sales, "--rename", "name=full_name", "-n")
    assert result.exit_code == 0
    assert "full_name" in result.stdout
    assert sales.read_text(encoding="utf-8") == original
    assert not (sales.parent / "sales.csv.bak").exists()


def test_edit_without_confirmation_writes_nothing(runner, sales):
    original = sales.read_text(encoding="utf-8")
    result = run(runner, "edit", sales, "--rename", "name=full_name")
    assert result.exit_code != 0
    assert sales.read_text(encoding="utf-8") == original


def test_edit_needs_something_to_do(runner, sales):
    result = run(runner, "edit", sales)
    assert result.exit_code != 0
    assert "--rename" in result.stderr


def test_edit_refuses_stdin(runner):
    result = run(runner, "edit", "-", "--from", "json", "--rename", "a=b", "-y", input='[{"a": 1}]')
    assert result.exit_code != 0
    assert "stdin" in result.stderr


def test_edit_refuses_an_output_that_is_the_source(runner, sales):
    result = run(runner, "edit", sales, "--rename", "name=x", "-o", sales, "-y")
    assert result.exit_code != 0
    assert "same file" in result.stderr


def test_edit_names_a_real_column_when_the_old_one_is_wrong(runner, sales):
    result = run(runner, "edit", sales, "--rename", "nmae=full_name", "-y")
    assert result.exit_code != 0
    assert "Did you mean: name?" in result.stderr


def test_edit_says_when_there_is_nothing_to_change(runner, sales):
    original = sales.read_text(encoding="utf-8")
    result = run(runner, "edit", sales, "--rename", "name=name", "-y")
    assert result.exit_code == 0
    assert "Nothing to change" in result.stderr
    assert sales.read_text(encoding="utf-8") == original


def test_edit_interactive_asks_for_every_name_in_turn(runner, sales):
    result = run(runner, "edit", sales, "-i", input="user_id\n\n\n\n\ny\n")
    assert result.exit_code == 0
    assert sales.read_text(encoding="utf-8").splitlines()[0] == "user_id,name,joined,zip,amount"


def test_edit_interactive_keeps_a_name_when_the_answer_is_blank(runner, sales):
    run(runner, "edit", sales, "-i", input="\n\n\n\n\ny\n")
    assert "Nothing to change" in run(runner, "edit", sales, "-i", input="\n\n\n\n\ny\n").stderr


def test_edit_interactive_can_be_declined(runner, sales):
    original = sales.read_text(encoding="utf-8")
    result = run(runner, "edit", sales, "-i", input="user_id\n\n\n\n\nn\n")
    assert result.exit_code != 0
    assert sales.read_text(encoding="utf-8") == original


def test_edit_interactive_suggests_snake_case_names(runner, book):
    run(runner, "edit", book, "-i", "--suggest", input="\n\n\ny\n")
    assert [cell.value for cell in openpyxl.load_workbook(book).active[1]] == ["id", "full_name", "hired"]


def test_edit_interactive_starts_from_a_rename_given_on_the_command_line(runner, sales):
    run(runner, "edit", sales, "-i", "--rename", "name=full_name", input="\n\n\n\n\ny\n")
    assert sales.read_text(encoding="utf-8").splitlines()[0] == "id,full_name,joined,zip,amount"


# ── sql: choosing columns ───────────────────────────────────────────


def test_sql_keeps_only_the_columns_asked_for(runner, flat):
    script = run(runner, "sql", flat, "-t", "users", "-k", "id", "-k", "active", "--sql", "-").stdout
    assert '"id"' in script and '"active"' in script
    assert "first_name" not in script


def test_sql_key_is_a_glob(runner, flat):
    script = run(runner, "sql", flat, "-t", "users", "-k", "*name*", "--sql", "-").stdout
    assert '"first_name"' in script
    assert '"active"' not in script


def test_sql_key_matches_the_original_spelling_too(runner, flat):
    script = run(runner, "sql", flat, "-t", "users", "-k", "First Name", "--sql", "-").stdout
    assert '"first_name"' in script
    assert '"id"' not in script


def test_sql_key_narrows_the_inserted_values(runner, flat):
    script = run(runner, "sql", flat, "-t", "users", "-k", "id", "--sql", "-").stdout
    insert = [line for line in script.splitlines() if line.strip().startswith("(")][0]
    assert insert.strip().startswith("(1)")


def test_sql_key_narrows_a_real_sqlite_table(runner, flat, tmp_path):
    db = tmp_path / "app.db"
    result = run(runner, "sql", flat, "-t", "users", "-k", "id", "-k", "age", "--db", db)
    assert result.exit_code == 0
    with sqlite3.connect(db) as connection:
        names = [row[1] for row in connection.execute("PRAGMA table_info(users)")]
    assert names == ["id", "age"]


def test_sql_key_matching_nothing_is_an_error(runner, flat):
    result = run(runner, "sql", flat, "-t", "users", "-k", "nope*", "--sql", "-")
    assert result.exit_code != 0
    assert "No field matched" in result.stderr


def test_sql_key_that_drops_the_primary_key_says_so(runner, flat):
    result = run(runner, "sql", flat, "-t", "users", "-k", "active", "--pk", "id", "--sql", "-")
    assert result.exit_code != 0
    assert "excluded by --key" in result.stderr


def test_sql_interactive_picks_columns_by_number(runner, flat):
    script = run(runner, "sql", flat, "-i", "--sql", "-", input="t\n1,2\n\n\n\ny\n").stdout
    assert '"id"' in script and '"first_name"' in script
    assert '"active"' not in script


def test_sql_interactive_picks_columns_by_name(runner, flat):
    script = run(runner, "sql", flat, "-i", "--sql", "-", input="t\nid,active\n\n\n\ny\n").stdout
    assert '"active"' in script
    assert '"first_name"' not in script


def test_sql_interactive_mixes_numbers_and_names(runner, flat):
    script = run(runner, "sql", flat, "-i", "--sql", "-", input="t\n1,active\n\n\n\ny\n").stdout
    assert '"id"' in script and '"active"' in script
    assert '"first_name"' not in script


def test_sql_interactive_keeps_every_column_when_the_answer_is_blank(runner, flat):
    script = run(runner, "sql", flat, "-i", "--sql", "-", input="t\n\n\n\n\ny\n").stdout
    assert '"id"' in script and '"first_name"' in script and '"active"' in script


def test_sql_interactive_lists_the_columns_with_their_numbers(runner, flat):
    result = run(runner, "sql", flat, "-i", "--sql", "-", input="t\n\n\n\n\ny\n")
    assert "1 id" in result.stderr and "2 first_name" in result.stderr


def test_sql_interactive_re_asks_after_an_unknown_column(runner, flat):
    result = run(runner, "sql", flat, "-i", "--sql", "-", input="t\nnope\nid\n\n\n\ny\n")
    assert result.exit_code == 0
    assert "No column called 'nope'" in result.stderr


def test_sql_interactive_re_asks_after_a_number_out_of_range(runner, flat):
    result = run(runner, "sql", flat, "-i", "--sql", "-", input="t\n99\nid\n\n\n\ny\n")
    assert result.exit_code == 0
    assert "99" in result.stderr


def test_sql_interactive_shows_how_many_columns_were_kept(runner, flat):
    result = run(runner, "sql", flat, "-i", "--sql", "-", input="t\nid,active\n\n\n\ny\n")
    assert "2 of 6" in result.stderr


def test_sql_interactive_suggests_keys_from_the_chosen_columns_only(runner, flat):
    result = run(runner, "sql", flat, "-i", "--sql", "-", input="t\nfirst_name,active\n\n\n\ny\n")
    line = [line for line in result.stderr.splitlines() if "unique and complete" in line][0]
    assert "id" not in line
