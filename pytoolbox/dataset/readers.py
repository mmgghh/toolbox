"""Reading JSON, CSV and Excel into records.

Each reader returns plain Python values plus the key order of the source, so
that everything downstream sees one shape regardless of where the data came
from. JSON already carries types; CSV carries none and gets them inferred;
Excel carries its own and keeps them.
"""

from __future__ import annotations

import csv
import io
import itertools
import json
from collections import deque
from collections.abc import Iterator
from pathlib import Path
from typing import Optional

from pytoolbox.core.console import plural
from pytoolbox.dataset import naming
from pytoolbox.dataset.errors import DataError
from pytoolbox.dataset.types import ValueType, parse_text, unify_all

#: Suffix -> reader kind, for detecting the input without being told.
SUFFIXES = {
    ".json": "json",
    ".ndjson": "json",
    ".jsonl": "json",
    ".csv": "csv",
    ".tsv": "csv",
    ".txt": "csv",
    ".xlsx": "excel",
    ".xlsm": "excel",
}

#: The kinds ``--from`` accepts.
KINDS = ("json", "csv", "excel")

#: Bytes of a CSV handed to the sniffer when no delimiter was given.
_SNIFF_BYTES = 8192


def detect_kind(path: Path) -> Optional[str]:
    """Guess the reader from a filename, or return None if the suffix is new."""
    return SUFFIXES.get(path.suffix.lower())


def read_text(path: Path, encoding: str = "utf-8", errors: str = "replace") -> str:
    """Read a file, or standard input when ``path`` is ``-``."""
    if str(path) == "-":
        import sys

        data = sys.stdin.buffer.read()
        return data.decode(encoding, errors)
    try:
        return path.read_text(encoding=encoding, errors=errors)
    except FileNotFoundError as exc:
        raise DataError(f"No such file: {path}") from exc
    except OSError as exc:
        raise DataError(f"Cannot read {path}: {exc}") from exc


def read_json(text: str) -> object:
    """Parse a JSON document, or a stream of one JSON value per line.

    Newline-delimited JSON is common enough as an export format that failing
    on it would be a papercut; it is tried only after a whole-document parse
    fails, so a pretty-printed file is never mistaken for it.
    """
    stripped = text.strip()
    if not stripped:
        raise DataError("The input is empty.")
    try:
        return json.loads(stripped)
    except json.JSONDecodeError as whole_document_error:
        lines = [line for line in stripped.splitlines() if line.strip()]
        if len(lines) > 1:
            try:
                return [json.loads(line) for line in lines]
            except json.JSONDecodeError:
                pass
        raise DataError(
            f"Not valid JSON: {whole_document_error.msg} "
            f"(line {whole_document_error.lineno}, column {whole_document_error.colno})"
        ) from whole_document_error


def resolve_delimiter(text: str, path: Optional[Path] = None, delimiter: Optional[str] = None) -> str:
    """The delimiter to read a file with: the given one, the suffix, or a sniff.

    Every caller resolves it here so that a file is never read with one
    delimiter and written back with another.
    """
    if delimiter:
        return delimiter
    if path is not None and path.suffix.lower() == ".tsv":
        return "\t"
    return _sniff_delimiter(text)


def read_csv(
    text: str,
    delimiter: Optional[str] = None,
    infer: bool = True,
) -> tuple[list[dict], list[str]]:
    """Read delimited text into records, inferring column types by default."""
    if not text.strip():
        raise DataError("The input is empty.")
    delimiter = resolve_delimiter(text, None, delimiter)

    reader = csv.reader(io.StringIO(text, newline=""), delimiter=delimiter)
    try:
        header = next(reader)
    except StopIteration as exc:  # pragma: no cover - guarded by the emptiness check
        raise DataError("The input has no header row.") from exc

    columns = _header_columns(header)
    raw_rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not raw_rows:
        raise DataError("The input has a header row but no data rows.")

    cells = _align(raw_rows, len(columns))
    values = _infer_columns(cells, len(columns)) if infer else _as_text(cells, len(columns))
    records = [
        {column: values[index][row] for index, column in enumerate(columns)}
        for row in range(len(cells))
    ]
    return records, columns


def stream_csv_rows(
    path: Path,
    encoding: str = "utf-8",
    errors: str = "replace",
    delimiter: Optional[str] = None,
) -> tuple[list[str], str, Iterator[list[str]], io.TextIOBase]:
    """Open a CSV/TSV for row-by-row reading: columns, delimiter, rows, and the stream to close.

    A real file is opened and read lazily, one row at a time, so a caller
    that stops early -- ``head`` after N rows -- never reads the rest of the
    file. Stdin can't be sniffed and then rewound the way a seekable file
    can, so it is buffered into memory up front, same as :func:`read_text`.
    """
    if str(path) == "-":
        import sys

        data = sys.stdin.buffer.read().decode(encoding, errors)
        if not data.strip():
            raise DataError("The input is empty.")
        stream: io.TextIOBase = io.StringIO(data)
    else:
        try:
            if path.stat().st_size == 0:
                raise DataError("The input is empty.")
        except FileNotFoundError as exc:
            raise DataError(f"No such file: {path}") from exc
        try:
            stream = open(path, "r", encoding=encoding, errors=errors, newline="")
        except OSError as exc:
            raise DataError(f"Cannot read {path}: {exc}") from exc

    used_delimiter = delimiter
    if not used_delimiter and str(path) != "-" and path.suffix.lower() == ".tsv":
        used_delimiter = "\t"
    if not used_delimiter:
        sample = stream.read(_SNIFF_BYTES)
        stream.seek(0)
        used_delimiter = _sniff_delimiter(sample)

    reader = csv.reader(stream, delimiter=used_delimiter)
    try:
        header = next(reader)
    except StopIteration as exc:
        stream.close()
        raise DataError("The input is empty.") from exc

    columns = _header_columns(header)
    rows = (row for row in reader if any(cell.strip() for cell in row))
    return columns, used_delimiter, rows, stream


def edge_csv_records(
    path: Path,
    encoding: str = "utf-8",
    errors: str = "replace",
    delimiter: Optional[str] = None,
    infer: bool = True,
    n: int = 10,
    from_end: bool = False,
    limit: Optional[int] = None,
) -> tuple[list[dict], list[str], list[str]]:
    """The first or last ``n`` data rows of a CSV/TSV, read without the rest.

    ``head`` (``from_end=False``) stops reading once ``n`` rows are
    collected. ``tail`` (``from_end=True``) must still stream every row --
    there is no way to find the true end of a CSV without it, since a quoted
    field can itself contain a newline -- but keeps only the last ``n`` rows
    in memory at once rather than the whole file. ``limit``, when given,
    bounds the rows considered before ``n`` is taken from either end.

    Raises like :func:`read_csv` when there turn out to be no data rows at
    all -- ``n`` only controls how many of them are kept, not whether the
    source itself is allowed to be empty.
    """
    columns, _, rows, stream = stream_csv_rows(path, encoding, errors, delimiter)
    try:
        notes: list[str] = []
        if limit is not None and limit >= 0:
            limited = list(itertools.islice(rows, limit))
            if next(rows, None) is not None:
                notes.append(f"Stopped at --limit {limit}.")
            rows = iter(limited)
        peeked = next(rows, None)
        if peeked is None:
            raise DataError("The input has a header row but no data rows.")
        rest = itertools.chain([peeked], rows)
        if from_end:
            raw_rows = list(deque(rest, maxlen=max(n, 0)))
        else:
            raw_rows = list(itertools.islice(rest, max(n, 0)))
    finally:
        stream.close()

    cells = _align(raw_rows, len(columns))
    values = _infer_columns(cells, len(columns)) if infer else _as_text(cells, len(columns))
    records = [
        {column: values[index][row] for index, column in enumerate(columns)}
        for row in range(len(cells))
    ]
    return records, columns, notes


def count_csv_rows(text: str, delimiter: str) -> int:
    """Count data rows in delimited text, without inferring types or building records.

    Used by ``count`` in place of :func:`read_csv`: the per-cell type
    inference :func:`read_csv` does is the dominant cost on a large file and
    is wasted work when only the row count is wanted.
    """
    if not text.strip():
        raise DataError("The input is empty.")
    reader = csv.reader(io.StringIO(text, newline=""), delimiter=delimiter)
    try:
        next(reader)
    except StopIteration as exc:  # pragma: no cover - guarded by the emptiness check
        raise DataError("The input has no header row.") from exc
    total = sum(1 for row in reader if any(cell.strip() for cell in row))
    if total == 0:
        raise DataError("The input has a header row but no data rows.")
    return total


def parse_json_for_count(text: str) -> tuple[object, Optional[int]]:
    """Parse JSON for counting: the document, or a cheap NDJSON line count.

    Mirrors :func:`read_json`'s fallback to newline-delimited JSON exactly,
    including that every line must parse for the fallback to be accepted --
    a document that is neither valid JSON as a whole nor line by line is
    still an error, not a count. Unlike :func:`read_json`, the parsed line
    values are not kept: counting only needs to know how many there are.
    """
    stripped = text.strip()
    if not stripped:
        raise DataError("The input is empty.")
    try:
        return json.loads(stripped), None
    except json.JSONDecodeError as whole_document_error:
        lines = [line for line in stripped.splitlines() if line.strip()]
        if len(lines) > 1:
            try:
                for line in lines:
                    json.loads(line)
            except json.JSONDecodeError:
                pass
            else:
                return None, len(lines)
        raise DataError(
            f"Not valid JSON: {whole_document_error.msg} "
            f"(line {whole_document_error.lineno}, column {whole_document_error.colno})"
        ) from whole_document_error


def edge_ndjson_lines(
    path: Path,
    encoding: str = "utf-8",
    errors: str = "replace",
    n: int = 10,
    from_end: bool = False,
    limit: Optional[int] = None,
) -> tuple[list[str], list[str]]:
    """The first or last ``n`` non-blank lines of a newline-delimited JSON file.

    Reads nothing beyond what it needs to: ``head`` stops after ``n`` lines
    without opening the rest of the file, and ``tail`` streams every line but
    keeps only the last ``n`` in memory. The caller parses each returned line
    -- this only ever handles text, and never raises on invalid JSON, only on
    there being no lines at all.
    """
    if str(path) == "-":
        import sys

        data = sys.stdin.buffer.read().decode(encoding, errors)
        lines = (line for line in data.splitlines() if line.strip())
        stream = None
    else:
        try:
            stream = open(path, "r", encoding=encoding, errors=errors)
        except FileNotFoundError as exc:
            raise DataError(f"No such file: {path}") from exc
        except OSError as exc:
            raise DataError(f"Cannot read {path}: {exc}") from exc
        lines = (line for line in stream if line.strip())

    try:
        notes: list[str] = []
        if limit is not None and limit >= 0:
            limited = list(itertools.islice(lines, limit))
            if next(lines, None) is not None:
                notes.append(f"Stopped at --limit {limit}.")
            lines = iter(limited)
        peeked = next(lines, None)
        if peeked is None:
            raise DataError("The input is empty.")
        rest = itertools.chain([peeked], lines)
        if from_end:
            return list(deque(rest, maxlen=max(n, 0))), notes
        return list(itertools.islice(rest, max(n, 0))), notes
    finally:
        if stream is not None:
            stream.close()


def count_excel_sheet(path: Path, sheet: Optional[str] = None) -> int:
    """Count the data rows of one worksheet, header row excluded.

    Like :func:`count_excel_sheets` but for a single, already-chosen sheet --
    used by ``count --sheet`` to avoid building a record per row.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DataError(
            "Reading .xlsx needs openpyxl. Install it with: pip install 'pytoolbox[excel]'"
        ) from exc

    if not path.exists():
        raise DataError(f"No such file: {path}")
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        if sheet is None:
            worksheet = workbook.active
        elif sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
        else:
            available = ", ".join(workbook.sheetnames)
            raise DataError(f"No sheet named {sheet!r}. This workbook has: {available}")
        rows = worksheet.iter_rows(values_only=True)
        next(rows, None)
        return sum(1 for row in rows if not _is_blank_row(row))
    finally:
        workbook.close()


def read_excel(
    path: Path,
    sheet: Optional[str] = None,
    infer: bool = True,
) -> tuple[list[dict], list[str], list[str]]:
    """Read a worksheet into records, keeping the cell types Excel recorded.

    Excel already distinguishes a number from a date from text, so there is
    nothing to infer; ``infer=False`` flattens every cell back to text for the
    cases where the recorded types are wrong.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DataError(
            "Reading .xlsx needs openpyxl. Install it with: pip install 'pytoolbox[excel]'"
        ) from exc

    if not path.exists():
        raise DataError(f"No such file: {path}")
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        notes = []
        if sheet is None:
            worksheet = workbook.active
            if len(workbook.sheetnames) > 1:
                notes.append(
                    f"Using --sheet {worksheet.title!r} "
                    f"({plural(len(workbook.sheetnames), 'sheet')} in this workbook)."
                )
        elif sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
        else:
            available = ", ".join(workbook.sheetnames)
            raise DataError(f"No sheet named {sheet!r}. This workbook has: {available}")

        rows = worksheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration as exc:
            raise DataError(f"Sheet {worksheet.title!r} is empty.") from exc

        columns = _header_columns(["" if cell is None else str(cell) for cell in header])
        records = []
        for row in rows:
            if _is_blank_row(row):
                continue
            record = {}
            for index, column in enumerate(columns):
                value = row[index] if index < len(row) else None
                if isinstance(value, str) and not value.strip():
                    value = None
                elif not infer and value is not None:
                    value = str(value)
                record[column] = value
            records.append(record)
    finally:
        workbook.close()

    if not records:
        raise DataError(f"Sheet {worksheet.title!r} has a header row but no data rows.")
    return records, columns, notes


def edge_excel_records(
    path: Path,
    sheet: Optional[str] = None,
    infer: bool = True,
    n: int = 10,
    from_end: bool = False,
    limit: Optional[int] = None,
) -> tuple[list[dict], list[str], list[str]]:
    """The first or last ``n`` data rows of one worksheet.

    Like :func:`edge_csv_records`: ``head`` stops after ``n`` rows without
    reading the rest of the sheet; ``tail`` streams every row -- openpyxl's
    read-only mode has no way to seek to the end -- but keeps only the last
    ``n`` in memory. Raises like :func:`read_excel` when the sheet turns out
    to have no data rows at all.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DataError(
            "Reading .xlsx needs openpyxl. Install it with: pip install 'pytoolbox[excel]'"
        ) from exc

    if not path.exists():
        raise DataError(f"No such file: {path}")
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        notes = []
        if sheet is None:
            worksheet = workbook.active
            if len(workbook.sheetnames) > 1:
                notes.append(
                    f"Using --sheet {worksheet.title!r} "
                    f"({plural(len(workbook.sheetnames), 'sheet')} in this workbook)."
                )
        elif sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
        else:
            available = ", ".join(workbook.sheetnames)
            raise DataError(f"No sheet named {sheet!r}. This workbook has: {available}")

        rows = worksheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration as exc:
            raise DataError(f"Sheet {worksheet.title!r} is empty.") from exc
        columns = _header_columns(["" if cell is None else str(cell) for cell in header])

        data_rows = (row for row in rows if not _is_blank_row(row))
        if limit is not None and limit >= 0:
            limited = list(itertools.islice(data_rows, limit))
            if next(data_rows, None) is not None:
                notes.append(f"Stopped at --limit {limit}.")
            data_rows = iter(limited)
        peeked = next(data_rows, None)
        if peeked is None:
            raise DataError(f"Sheet {worksheet.title!r} has a header row but no data rows.")
        rest = itertools.chain([peeked], data_rows)
        selected = deque(rest, maxlen=max(n, 0)) if from_end else itertools.islice(rest, max(n, 0))

        records = []
        for row in selected:
            record = {}
            for index, column in enumerate(columns):
                value = row[index] if index < len(row) else None
                if isinstance(value, str) and not value.strip():
                    value = None
                elif not infer and value is not None:
                    value = str(value)
                record[column] = value
            records.append(record)
    finally:
        workbook.close()

    return records, columns, notes


def _is_blank_row(row: tuple) -> bool:
    """True when every cell is empty, the way a spreadsheet's blank rows are."""
    return all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row)


def list_excel_sheets(path: Path) -> list[str]:
    """The name of every sheet in a workbook, in workbook order."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DataError(
            "Reading .xlsx needs openpyxl. Install it with: pip install 'pytoolbox[excel]'"
        ) from exc

    if not path.exists():
        raise DataError(f"No such file: {path}")
    workbook = load_workbook(filename=str(path), read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def count_excel_sheets(path: Path) -> dict[str, int]:
    """Count the data rows in every sheet of a workbook, header row excluded.

    Unlike :func:`read_excel`, an empty sheet counts as zero instead of
    raising -- counting is meant to survey a workbook, not load it.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DataError(
            "Reading .xlsx needs openpyxl. Install it with: pip install 'pytoolbox[excel]'"
        ) from exc

    if not path.exists():
        raise DataError(f"No such file: {path}")
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        counts = {}
        for name in workbook.sheetnames:
            rows = workbook[name].iter_rows(values_only=True)
            next(rows, None)
            counts[name] = sum(1 for row in rows if not _is_blank_row(row))
        return counts
    finally:
        workbook.close()


def read_excel_headers(path: Path) -> dict[str, list[str]]:
    """The column headers of every sheet in a workbook.

    Like :func:`count_excel_sheets`, a sheet with no header row is given an
    empty list instead of raising -- this is meant to survey a workbook, not
    load it.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DataError(
            "Reading .xlsx needs openpyxl. Install it with: pip install 'pytoolbox[excel]'"
        ) from exc

    if not path.exists():
        raise DataError(f"No such file: {path}")
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        headers = {}
        for name in workbook.sheetnames:
            header = next(workbook[name].iter_rows(values_only=True), None)
            headers[name] = (
                _header_columns(["" if cell is None else str(cell) for cell in header])
                if header
                else []
            )
        return headers
    finally:
        workbook.close()


def _sniff_delimiter(text: str) -> str:
    """Work out the delimiter, falling back to a comma when unsure."""
    sample = text[:_SNIFF_BYTES]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        return ","


def _header_columns(header: list[str]) -> list[str]:
    """Name every header cell, filling in blanks and breaking duplicates."""
    named = [cell.strip() or f"{naming.FALLBACK}_{index + 1}" for index, cell in enumerate(header)]
    seen: dict[str, int] = {}
    columns = []
    for name in named:
        count = seen.get(name, 0) + 1
        seen[name] = count
        columns.append(name if count == 1 else f"{name}_{count}")
    return columns


def _align(rows: list[list[str]], width: int) -> list[list[str]]:
    """Pad short rows and drop the overflow of long ones."""
    return [(row + [""] * width)[:width] for row in rows]


def _as_text(cells: list[list[str]], width: int) -> list[list[object]]:
    """Keep every cell as text, with blanks as nulls."""
    return [[(row[index].strip() or None) for row in cells] for index in range(width)]


def _infer_columns(cells: list[list[str]], width: int) -> list[list[object]]:
    """Type each column as a whole; one unparseable cell keeps it as text.

    Deciding per column rather than per cell is what stops a mostly-numeric
    column with one ``n/a`` in it from becoming a column of mixed types.
    """
    columns: list[list[object]] = []
    for index in range(width):
        raw = [row[index] for row in cells]
        parsed = [parse_text(cell) for cell in raw]
        resolved = unify_all(value_type for value_type, _ in parsed)
        if resolved in (ValueType.STR, ValueType.NULL):
            columns.append([cell.strip() or None for cell in raw])
        else:
            columns.append([value for _, value in parsed])
    return columns
