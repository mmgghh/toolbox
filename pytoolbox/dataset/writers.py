"""Writing renamed names back into a file, and nothing else.

This module touches the disk and makes no decisions: it is handed a plan and
carries it out. Each format is rewritten as narrowly as it can be -- a CSV
keeps every data row byte for byte, a workbook keeps its formulas and styles,
a JSON document keeps its key order and its indentation -- because a rename
that reformats the file is a rename nobody can review.

The transforms for the text formats are pure functions of the text, so they
can be tested without a filesystem; only :func:`apply` writes anything.
"""

from __future__ import annotations

import csv
import io
import json
import os
import shutil
import tempfile
from pathlib import Path
from typing import Optional

from pytoolbox.dataset import readers
from pytoolbox.dataset.edit import RenamePlan
from pytoolbox.dataset.errors import DataError

#: Suffix of the copy taken before an in-place edit.
BACKUP_SUFFIX = ".bak"


def rewrite_csv(text: str, plan: RenamePlan, delimiter: str) -> str:
    """Return ``text`` with only its header line rewritten.

    The header record is re-serialized and spliced in front of the original
    remaining text, so quoting, spacing and line endings in the data rows are
    exactly what they were.
    """
    reader = csv.reader(io.StringIO(text, newline=""), delimiter=delimiter)
    try:
        header = next(reader)
    except StopIteration as exc:  # pragma: no cover - an empty file never gets here
        raise DataError("The input has no header row.") from exc

    for item in plan.renames:
        if item.index < len(header):
            header[item.index] = item.new

    lines = text.splitlines(keepends=True)
    # A quoted header field can span physical lines; line_num counts how many.
    consumed = max(reader.line_num, 1)
    last = lines[consumed - 1]
    ending = last[len(last.rstrip("\r\n")):]

    buffer = io.StringIO()
    csv.writer(buffer, delimiter=delimiter, lineterminator="").writerow(header)
    return buffer.getvalue() + ending + "".join(lines[consumed:])


def rewrite_json(text: str, plan: RenamePlan, root: str = "") -> str:
    """Return ``text`` with the planned keys renamed and nothing else moved."""
    from pytoolbox.dataset import sources

    mapping = {item.old: item.new for item in plan.renames}
    stripped = text.strip()
    trailing = text[len(text.rstrip("\n")):]

    try:
        document = json.loads(stripped)
    except json.JSONDecodeError:
        return _rewrite_ndjson(text, mapping)

    if root:
        segments = root.split(".")
        parent = sources.resolve_path(document, ".".join(segments[:-1])) if len(segments) > 1 else document
        parent[segments[-1]] = _renamed(parent[segments[-1]], mapping)
    else:
        document = _renamed(document, mapping)

    indent = _indent_of(stripped)
    return json.dumps(document, indent=indent, ensure_ascii=False) + trailing


def rewrite_excel(source: Path, target: Path, plan: RenamePlan, sheet: Optional[str] = None) -> None:
    """Copy ``source`` to ``target`` with its header cells renamed.

    The workbook is opened for writing rather than read-only, and only the
    header row's values are assigned, so formulas, styles, column widths and
    every other sheet come through untouched.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DataError(
            "Editing .xlsx needs openpyxl. Install it with: pip install 'pytoolbox[excel]'"
        ) from exc

    workbook = load_workbook(filename=str(source), keep_vba=source.suffix.lower() == ".xlsm")
    try:
        if sheet is None:
            worksheet = workbook.active
        elif sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
        else:
            available = ", ".join(workbook.sheetnames)
            raise DataError(f"No sheet named {sheet!r}. This workbook has: {available}")
        for item in plan.renames:
            worksheet.cell(row=1, column=item.index + 1).value = item.new
        workbook.save(str(target))
    finally:
        workbook.close()


def apply(
    plan: RenamePlan,
    path: Path,
    kind: str,
    target: Optional[Path] = None,
    sheet: Optional[str] = None,
    delimiter: str = ",",
    encoding: str = "utf-8",
    errors: str = "replace",
    root: str = "",
    backup: bool = True,
) -> Optional[Path]:
    """Carry out ``plan``, returning the backup path when one was written.

    The new file is written beside its destination and moved into place, so an
    interrupted run leaves either the old file or the new one, never half of
    either.
    """
    target = target or path
    backup_path = None
    if backup and target == path:
        backup_path = path.with_name(path.name + BACKUP_SUFFIX)
        shutil.copy2(path, backup_path)

    temporary = _temp_path(target)
    try:
        if kind == "excel":
            rewrite_excel(path, temporary, plan, sheet)
        else:
            text = readers.read_text(path, encoding=encoding, errors=errors)
            if kind == "csv":
                rewritten = rewrite_csv(text, plan, delimiter)
            elif kind == "json":
                rewritten = rewrite_json(text, plan, root)
            else:  # pragma: no cover - guarded by the reader's own check
                raise DataError(f"Cannot edit {kind!r} files.")
            temporary.write_text(rewritten, encoding=encoding, errors=errors)
        os.replace(temporary, target)
    except BaseException:
        temporary.unlink(missing_ok=True)
        raise
    return backup_path


def _temp_path(target: Path) -> Path:
    """A scratch path beside ``target``, so the move into place is atomic."""
    target.parent.mkdir(parents=True, exist_ok=True)
    handle, name = tempfile.mkstemp(dir=str(target.parent), prefix=f".{target.name}.", suffix=".tmp")
    os.close(handle)
    return Path(name)


def _renamed(selected: object, mapping: dict[str, str]) -> object:
    """Rename the top-level keys of a record, or of every record in a list."""
    if isinstance(selected, dict):
        return {mapping.get(key, key): value for key, value in selected.items()}
    if isinstance(selected, list):
        return [_renamed(item, mapping) if isinstance(item, dict) else item for item in selected]
    return selected


def _rewrite_ndjson(text: str, mapping: dict[str, str]) -> str:
    """Rewrite one JSON value per line, keeping that shape."""
    out = []
    for line in text.splitlines(keepends=True):
        body = line.rstrip("\r\n")
        ending = line[len(body):]
        if not body.strip():
            out.append(line)
            continue
        out.append(json.dumps(_renamed(json.loads(body), mapping), ensure_ascii=False) + ending)
    return "".join(out)


def _indent_of(text: str) -> Optional[object]:
    """The indentation the document was written with, or None if it was flat."""
    for line in text.splitlines()[1:]:
        body = line.lstrip(" \t")
        if not body or len(body) == len(line):
            continue
        lead = line[: len(line) - len(body)]
        return "\t" if lead[0] == "\t" else len(lead)
    return None
